"""
Saccharomyces cerevisiae Fermentation Simulator

Based on:
Lei, F., Rotboll, M., & Jorgensen, S. B. (2001).
A biochemically structured model for Saccharomyces cerevisiae.
Journal of Biotechnology, 88(3), 205-221.
https://doi.org/10.1016/s0168-1656(01)00269-3

Run with:
    streamlit run fermentation_app.py
"""

from __future__ import annotations

import csv
import io
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy.integrate import solve_ivp
import streamlit as st
from openpyxl import load_workbook, Workbook


# -----------------------------
# Model parameters (kinetic constants)
# -----------------------------
PARAMS = {
    "k1_h": 0.584, "M1_h": 0.0116,
    "k1_L": 1.43, "M1_L": 0.94,
    "k1_e": 47.1, "M1_e": 0.12,
    "M1_i": 14.2,
    "k2": 0.501, "M2": 2e-5, "M2_i": 0.101,
    "k3": 5.81, "M3": 5e-7,
    "k4": 4.80, "M4": 2.64e-4,
    "k5": 0.0104, "M5": 0.0102,
    "k5_e": 0.775, "M5_e": 0.10, "M5_i": 440,
    "k6": 2.82, "M6": 0.034,
    "k6_r": 0.0125, "M6_e": 0.057,
    "k7": 1.203, "M7": 0.0101,
    "k8": 0.589,
    "k9": 0.008, "M9": 1e-6,
    "k9_e": 0.0751, "M9_e": 13, "M9_i": 25,
    "k9_c": 3.99e-3,
    "k10": 0.392, "M10": 2.3e-3,
    "k10_e": 3.39e-3, "M10_e": 1.8e-3,
    "k11": 0.02,
}

STATE_LABELS = {
    0: "Glucose",
    1: "Pyruvate",
    2: "Acetaldehyde",
    3: "Acetate",
    4: "Ethanol",
    5: "Biomass (X)",
    6: "Active biomass fraction (X_A)",
    7: "Alcohol dehydrogenase (X_Acdh)",
    8: "qO2",
    9: "qCO2",
    10: "RQO",
}

STATE_COLORS = {
    0: "#4C78A8",
    1: "#F58518",
    2: "#E45756",
    3: "#72B7B2",
    4: "#54A24B",
    5: "#Eeca3b",
    6: "#B279a2",
    7: "#9D7660",
    8: "#FF9DA6",
    9: "#BAB0AC",
    10: "#EDC948",
}

SIM_HEADERS = [
    "Time (h)", "Glucose", "Pyruvate", "Acetaldehyde",
    "Acetate", "Ethanol", "Biomass", "X_A", "X_Acdh",
    "qO2", "qCO2", "RQO",
]


# -----------------------------
# ODE system (translated from MATLAB)
# -----------------------------
def ode_system(t, Y, Sf, Ds, Dw):
    p = PARAMS
    Glu, Pyr, Acetaldehyde, Acetate, EtOH, X, X_A, X_Acdh = Y
    D = Ds + Dw

    r1 = (p["k1_L"] * Glu * X_A / (Glu + p["M1_L"])
          + p["k1_h"] * Glu * X_A / (Glu + p["M1_h"])
          + p["k1_e"] * Glu * X_A * Acetaldehyde
          / (Glu * (p["M1_i"] * Acetaldehyde + 1) + p["M1_e"]))

    r2 = (p["k2"] * Pyr / (Pyr + p["M2"])
          * 1.0 / (p["M2_i"] * Glu + 1) * X_A)

    r3 = p["k3"] * Pyr**4 / (Pyr**4 + p["M3"]) * X_A

    r4 = (p["k4"] * Acetaldehyde / (p["M4"] + Acetaldehyde)
          * X_A * X_Acdh)

    r5 = (p["k5"] * Acetate * X_A / (Acetate + p["M5"])
          + p["k5_e"] * Acetate / (Acetate + p["M5_e"])
          * 1.0 / (1 + p["M5_i"] * Glu) * X_A)

    r6 = (p["k6"] * (Acetaldehyde - p["k6_r"] * EtOH) * X_A
          / (Acetaldehyde + p["M6"] + p["M6_e"] * EtOH))

    r7 = p["k7"] * Glu / (Glu + p["M7"]) * X_A

    r8 = (p["k8"] * Acetate / (Acetate + p["M5_e"])
          * 1.0 / (1 + p["M5_i"] * Glu) * X_A)

    r9 = ((p["k9"] * Glu / (Glu + p["M9"])
           + p["k9_e"] * EtOH / (EtOH + p["M9_e"]))
          * X_A / (1 + p["M9_i"] * Glu)
          + p["k9_c"] * Glu / (Glu + p["M9"]) * X_A)

    r10 = (p["k10"] * Glu / (Glu + p["M10"]) * X_A
           + p["k10_e"] * EtOH / (EtOH + p["M10_e"]) * X_A)

    r11 = p["k11"] * X_Acdh

    dGLu = -(r1 + r7) * X + (Sf * Ds - Glu * D)
    dPyr = (0.978 * r1 - r2 - r3) * X - Pyr * D
    dAcetaldehyde = (0.5 * r3 - r4 - r6) * X - Acetaldehyde * D
    dAcetate = (1.363 * r4 - r5 - r8) * X - Acetate * D
    dEtOH = 1.045 * r6 * X - EtOH * D
    dX = (0.732 * r7 + 0.619 * r8 - D) * X
    dX_A = (0.732 * r7 + 0.619 * r8 - r9 - r10
            - (0.732 * r7 + 0.619 * r8) * X_A)
    dX_Acdh = r9 - r11 - (0.732 * r7 + 0.619 * r8) * X_Acdh

    return [dGLu, dPyr, dAcetaldehyde, dAcetate,
            dEtOH, dX, dX_A, dX_Acdh]


def compute_auxiliary(Y):
    """Compute qO2, qCO2, RQO from state matrix. ALL rates must be recomputed."""
    p = PARAMS
    Glu = Y[:, 0]
    Pyr = Y[:, 1]
    Acetaldehyde = Y[:, 2]
    Acetate = Y[:, 3]
    EtOH = Y[:, 4]
    X_A = Y[:, 6]
    X_Acdh = Y[:, 7]

    r1 = (p["k1_L"] * Glu * X_A / (Glu + p["M1_L"])
          + p["k1_h"] * Glu * X_A / (Glu + p["M1_h"])
          + p["k1_e"] * Glu * X_A * Acetaldehyde
          / (Glu * (p["M1_i"] * Acetaldehyde + 1) + p["M1_e"]))

    r2 = (p["k2"] * Pyr / (Pyr + p["M2"])
          * 1.0 / (p["M2_i"] * Glu + 1) * X_A)

    r3 = p["k3"] * Pyr**4 / (Pyr**4 + p["M3"]) * X_A

    r4 = (p["k4"] * Acetaldehyde / (p["M4"] + Acetaldehyde)
          * X_A * X_Acdh)

    r5 = (p["k5"] * Acetate * X_A / (Acetate + p["M5"])
          + p["k5_e"] * Acetate / (Acetate + p["M5_e"])
          * 1.0 / (1 + p["M5_i"] * Glu) * X_A)

    r6 = (p["k6"] * (Acetaldehyde - p["k6_r"] * EtOH) * X_A
          / (Acetaldehyde + p["M6"] + p["M6_e"] * EtOH))

    r7 = p["k7"] * Glu / (Glu + p["M7"]) * X_A

    r8 = (p["k8"] * Acetate / (Acetate + p["M5_e"])
          * 1.0 / (1 + p["M5_i"] * Glu) * X_A)

    qO2 = (1.0 / 32.0 * 1e3
           * (0.178 * r1 + 0.908 * r2 + 0.363 * r4
              + 1.066 * r5 - 0.363 * r6 + 0.063 * r7 + 0.214 * r8))

    qCO2 = (1.0 / 44.01 * 1e3
            * (1.499 * r2 + 0.5 * r3 + 1.466 * r5
               + 0.127 * r7 + 0.325 * r8))

    RQO = np.divide(qCO2, qO2, out=np.zeros_like(qCO2), where=qO2 != 0)

    return qO2, qCO2, RQO


# -----------------------------
# Metrics
# -----------------------------
def rmse(y_true, y_pred):
    return np.sqrt(np.mean((y_true - y_pred) ** 2))


def mape(y_true, y_pred):
    mask = y_true != 0
    if not np.any(mask):
        return np.nan
    return np.mean(np.abs((y_true[mask] - y_pred[mask]) / y_true[mask])) * 100


# -----------------------------
# Excel I/O helpers (no pandas)
# -----------------------------
def read_excel(uploaded_file):
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel file must have at least a header and one data row.")
    data = []
    for row in rows[1:]:
        try:
            t = float(row[0])
            glu = float(row[1])
            bio = float(row[2])
            data.append((t, glu, bio))
        except (TypeError, ValueError):
            continue
    if not data:
        raise ValueError("No valid numeric data found in the Excel file.")
    return np.array([d[0] for d in data]), np.array([d[1] for d in data]), np.array([d[2] for d in data])


def build_csv_buffer(time_sim, Y_sim, qO2, qCO2, RQO):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(SIM_HEADERS)
    for i in range(len(time_sim)):
        writer.writerow([
            time_sim[i], Y_sim[i, 0], Y_sim[i, 1], Y_sim[i, 2],
            Y_sim[i, 3], Y_sim[i, 4], Y_sim[i, 5], Y_sim[i, 6],
            Y_sim[i, 7], qO2[i], qCO2[i], RQO[i],
        ])
    buf.seek(0)
    return buf.getvalue()


def build_excel_buffer(time_sim, Y_sim, qO2, qCO2, RQO, Sf, Ds, Dw,
                       rmse_glu, rmse_bio, mape_glu, mape_bio):
    buf = io.BytesIO()
    wb = Workbook()

    ws = wb.active
    ws.title = "Results"
    ws.append(SIM_HEADERS)
    for i in range(len(time_sim)):
        ws.append([
            time_sim[i], Y_sim[i, 0], Y_sim[i, 1], Y_sim[i, 2],
            Y_sim[i, 3], Y_sim[i, 4], Y_sim[i, 5], Y_sim[i, 6],
            Y_sim[i, 7], qO2[i], qCO2[i], RQO[i],
        ])

    ws2 = wb.create_sheet(title="Parameters")
    ws2.append(["Parameter", "Value"])
    ws2.append(["Sf (g/L)", Sf])
    ws2.append(["Ds (1/h)", Ds])
    ws2.append(["Dw (1/h)", Dw])
    ws2.append(["RMSE_Glucose", rmse_glu])
    ws2.append(["RMSE_Biomass", rmse_bio])
    ws2.append(["MAPE_Glucose (%)", mape_glu])
    ws2.append(["MAPE_Biomass (%)", mape_bio])

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# -----------------------------
# Plotting
# -----------------------------
def make_plot(time_sim, y_sim, time_exp, y_exp, state_idx, label):
    fig = go.Figure()

    fig.add_trace(go.Scatter(
        x=time_sim, y=y_sim,
        mode="lines",
        name=f"Simulated {label}",
        line=dict(color=STATE_COLORS[state_idx], width=2.5),
    ))

    if y_exp is not None and state_idx in (0, 5):
        fig.add_trace(go.Scatter(
            x=time_exp, y=y_exp,
            mode="markers",
            name=f"Experimental {label}",
            marker=dict(color=STATE_COLORS[state_idx], size=10,
                        symbol="diamond", line=dict(color="black", width=1)),
        ))

    fig.update_layout(
        title=dict(text=f"<b>{label}</b>", x=0.5),
        xaxis_title="Time (h)",
        yaxis_title="Concentration",
        template="plotly_white",
        height=500,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=60, r=40, t=80, b=60),
    )
    return fig


# -----------------------------
# Streamlit app
# -----------------------------
def main():
    st.set_page_config(
        page_title="S. cerevisiae Fermentation Simulator",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.title("S. cerevisiae Fermentation Simulator")
    st.caption(
        "Biochemically structured model based on Lei, Rotboll & Jorgensen (2001). "
        "Upload your experimental data, run the simulation, and compare results."
    )

    # ============================================================
    # SIDEBAR - MODIFY THESE LINES WITH YOUR LOGO PATHS
    # ============================================================
    with st.sidebar:
        # --- TOP LOGO ---
        # CHANGE THIS LINE: Replace the path with your top logo file
        st.image(
            "GIANT_LOGO-NO_BACKGROUND_HD.png",
            use_container_width=True,
        )
        # To control width/height manually, comment out the line above and use:
        # st.image("YOUR_PATH_HERE.png", width=200)
        # or
        # st.image("YOUR_PATH_HERE.png", width=200, output_format="PNG")

        st.header("Data Upload")
        uploaded_file = st.file_uploader(
            "Upload Excel file (Time | Glucose | Biomass)",
            type=["xlsx", "xls"],
        )

        st.header("Operating Parameters")
        Sf = st.number_input(
            "Glucose feed concentration, Sf (g/L)",
            min_value=0.0, max_value=500.0, value=0.0, step=1.0,
            help="Substrate concentration in the feed stream",
        )
        Ds = st.number_input(
            "Substrate dilution rate, Ds (1/h)",
            min_value=0.0, max_value=2.0, value=0.0, step=0.01,
            help="Dilution rate due to substrate feed",
        )
        Dw = st.number_input(
            "Washout dilution rate, Dw (1/h)",
            min_value=0.0, max_value=2.0, value=0.0, step=0.01,
            help="Dilution rate due to washout",
        )

        st.header("Solver Settings")
        solver_method = st.selectbox(
            "Integration method",
            ["LSODA (auto-stiff)", "BDF (stiff)", "RK45 (non-stiff)"],
            index=0,
            help="LSODA automatically switches between stiff and non-stiff. "
                 "Use BDF if LSODA is still slow. Avoid RK45 for this model.",
        )
        solver_map = {
            "LSODA (auto-stiff)": "LSODA",
            "BDF (stiff)": "BDF",
            "RK45 (non-stiff)": "RK45",
        }
        method = solver_map[solver_method]

        st.header("Plot Controls")
        plot_choice = st.selectbox(
            "Select variable to plot",
            options=list(STATE_LABELS.keys()),
            format_func=lambda k: STATE_LABELS[k],
            index=0,
        )
        plot_btn = st.button("Plot", use_container_width=True, type="primary")

        st.header("Export")
        st.info("Run the simulation first to enable download.")

        # --- BOTTOM LOGO ---
        # CHANGE THIS LINE: Replace the path with your bottom logo file
        st.image(
            "FOTO PERFIL SIN FONDO.png",
            use_container_width=True,
        )
        # To control width/height manually, comment out the line above and use:
        # st.image("YOUR_PATH_HERE.png", width=150)
    # ============================================================

    # Main area
    if uploaded_file is None:
        st.info("Please upload an Excel file with columns: Time | Glucose | Biomass to begin.")
        st.markdown("""
        ### Expected file format
        | Time (h) | Glucose (g/L) | Biomass (g/L) |
        |----------|---------------|---------------|
        | 0.0      | 20.0          | 0.5           |
        | 2.0      | 18.5          | 0.7           |
        | ...      | ...           | ...           |

        - **Column 1**: Time in hours
        - **Column 2**: Glucose concentration
        - **Column 3**: Biomass concentration

        All other model states start at 1e-3 and are integrated numerically.
        """)
        st.stop()

    # Read data
    try:
        time_exp, glu_exp, bio_exp = read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        st.stop()

    # Initial conditions
    Y0 = np.ones(8) * 1e-3
    Y0[0] = glu_exp[0]
    Y0[5] = bio_exp[0]

    t_span = (float(time_exp[0]), float(time_exp[-1]))
    t_eval = time_exp

    # Solve ODE with progress indicator
    with st.status("Integrating ODE system...", expanded=True) as status:
        st.write(f"Solver: **{method}** | Time span: {t_span[0]:.2f} to {t_span[1]:.2f} h")
        st.write(f"Eval points: **{len(t_eval)}** | Initial Glu: {Y0[0]:.4f} | Initial X: {Y0[5]:.4f}")

        sol = solve_ivp(
            fun=lambda t, y: ode_system(t, y, Sf, Ds, Dw),
            t_span=t_span,
            y0=Y0,
            t_eval=t_eval,
            method=method,
            dense_output=True,
        )

        if sol.success:
            status.update(label=f"Integration complete! ({sol.nfev:,} function evaluations)", state="complete")
        else:
            status.update(label=f"Integration failed: {sol.message}", state="error")

    if not sol.success:
        st.error(f"ODE solver failed: {sol.message}")
        st.stop()

    time_sim = sol.t
    Y_sim = sol.y.T

    # Compute auxiliary variables
    qO2, qCO2, RQO = compute_auxiliary(Y_sim)

    # Metrics
    rmse_glu = rmse(glu_exp, Y_sim[:, 0])
    rmse_bio = rmse(bio_exp, Y_sim[:, 5])
    mape_glu = mape(glu_exp, Y_sim[:, 0])
    mape_bio = mape(bio_exp, Y_sim[:, 5])

    # Display metrics
    st.subheader("Error Metrics")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("RMSE Glucose", f"{rmse_glu:.4f} g/L")
    m2.metric("RMSE Biomass", f"{rmse_bio:.4f} g/L")
    m3.metric("MAPE Glucose", f"{mape_glu:.2f} %" if not np.isnan(mape_glu) else "N/A")
    m4.metric("MAPE Biomass", f"{mape_bio:.2f} %" if not np.isnan(mape_bio) else "N/A")

    # Plotting
    if plot_btn:
        st.subheader("Simulation Results")

        if plot_choice in (0, 5):
            y_exp = glu_exp if plot_choice == 0 else bio_exp
        else:
            y_exp = None

        if plot_choice < 8:
            y_sim_plot = Y_sim[:, plot_choice]
        elif plot_choice == 8:
            y_sim_plot = qO2
        elif plot_choice == 9:
            y_sim_plot = qCO2
        else:
            y_sim_plot = RQO

        fig = make_plot(
            time_sim, y_sim_plot,
            time_exp, y_exp,
            plot_choice,
            STATE_LABELS[plot_choice],
        )
        st.plotly_chart(fig, use_container_width=True)

        # Comparison subplot for Glucose and Biomass
        if plot_choice in (0, 5):
            st.markdown("---")
            st.subheader("Glucose & Biomass Comparison")
            fig2 = make_subplots(
                rows=1, cols=2,
                subplot_titles=("Glucose", "Biomass"),
            )

            fig2.add_trace(go.Scatter(
                x=time_sim, y=Y_sim[:, 0],
                mode="lines", name="Simulated Glucose",
                line=dict(color=STATE_COLORS[0], width=2),
            ), row=1, col=1)
            fig2.add_trace(go.Scatter(
                x=time_exp, y=glu_exp,
                mode="markers", name="Experimental Glucose",
                marker=dict(color=STATE_COLORS[0], size=8, symbol="diamond"),
            ), row=1, col=1)

            fig2.add_trace(go.Scatter(
                x=time_sim, y=Y_sim[:, 5],
                mode="lines", name="Simulated Biomass",
                line=dict(color=STATE_COLORS[5], width=2),
            ), row=1, col=2)
            fig2.add_trace(go.Scatter(
                x=time_exp, y=bio_exp,
                mode="markers", name="Experimental Biomass",
                marker=dict(color=STATE_COLORS[5], size=8, symbol="diamond"),
            ), row=1, col=2)

            fig2.update_layout(
                template="plotly_white",
                height=450,
                showlegend=False,
                margin=dict(l=60, r=40, t=60, b=60),
            )
            fig2.update_xaxes(title_text="Time (h)")
            fig2.update_yaxes(title_text="Concentration (g/L)")
            st.plotly_chart(fig2, use_container_width=True)

    # Export
    st.subheader("Export Results")
    col_dl1, col_dl2 = st.columns(2)

    with col_dl1:
        csv_data = build_csv_buffer(time_sim, Y_sim, qO2, qCO2, RQO)
        st.download_button(
            label="Download CSV",
            data=csv_data,
            file_name="fermentation_simulation_results.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_dl2:
        excel_data = build_excel_buffer(
            time_sim, Y_sim, qO2, qCO2, RQO,
            Sf, Ds, Dw, rmse_glu, rmse_bio, mape_glu, mape_bio,
        )
        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name="fermentation_simulation_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # Citation
    st.markdown("---")
    st.markdown(
        """
        <div style="font-size: 0.85em; color: #666; text-align: center; padding: 1em;">
        <b>Reference:</b> Lei, F., Rotboll, M., & Jorgensen, S. B. (2001).
        A biochemically structured model for <i>Saccharomyces cerevisiae</i>.
        <i>Journal of Biotechnology</i>, 88(3), 205-221.
        <a href="https://doi.org/10.1016/s0168-1656(01)00269-3" target="_blank">
        https://doi.org/10.1016/s0168-1656(01)00269-3
        </a>
        </div>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
